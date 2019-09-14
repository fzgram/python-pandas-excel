#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import os
import sys
import csv
import re
import matplotlib.pyplot as plt
import openpyxl


# In[2]:


def is_interactive():
    import __main__ as main
    return not hasattr(main, '__file__')


# ## Sheet 1 - all runs

# In[3]:


def add_all_runs_sheet(c_df, writer):
    c_df.to_excel(writer, sheet_name='All runs', index=False)    


# ## Sheet 2 - only non-outliers from all runs

# In[4]:


def add_runs_without_outliers(c_df, writer, outlier_threshold=120):
    no_outliers_df = c_df[c_df['AR_fps'] >= outlier_threshold]
    no_outliers_df.to_excel(writer, sheet_name='Runs without outliers', index=False)
    return no_outliers_df


# ## Sheet 3 - only outliers from all runs

# In[5]:


def add_outlier_runs(c_df, writer, outlier_threshold=120):
    with_outliers_df = c_df[c_df['AR_fps'] < outlier_threshold]
    with_outliers_df.to_excel(writer, sheet_name='Outlier runs', index=False)
    return with_outliers_df


# ## Sheet 4 - add statistics (min, max, stddev, etc.) for FPS (runs without outliers)

# In[6]:


def add_statistics(c_df, writer):
    sliced_df = c_df[['run', 'AR_fps']].copy()
    desc_df = sliced_df.describe()
    median_series = sliced_df.median()
    
    fps_stat_df = pd.DataFrame(columns=['Min', 'Max', 'Average', 'Median', 'Std Deviation'])
         
    fps_stat_df['Min'] = [desc_df["AR_fps"]["min"]]
    fps_stat_df['Max'] = [desc_df["AR_fps"]["max"]]
    fps_stat_df['Average'] = [desc_df["AR_fps"]["mean"]]
    fps_stat_df['Median'] = [median_series["AR_fps"]]
    fps_stat_df['Std Deviation'] = [desc_df["AR_fps"]["std"]]

    # Round to 4 decimals
    fps_stat_df = fps_stat_df.round(4)
    
    # Write the sliced_df to excel
    sliced_df.to_excel(writer, sheet_name='Statistics', index=False)
    
    #get a pointer to the same sheet to write other dfs and text to the same sheet
    curr_sheet = writer.sheets['Statistics']
    
    # Write text and fps_stat_df
    #curr_sheet.write(1, 4, "Statistics, # of Frames Delay")
    curr_sheet['E2'] = "Statistics of FPS values"
    fps_stat_df.to_excel(writer, startrow=2, startcol=4, sheet_name='Statistics', index=False)
    
    return fps_stat_df


# ## Sheet 5 - Analyze AR_fps column from all runs

# In[7]:


def fps_all_analysis(c_df, writer):
    curr_row = 0
    fps_col_series = c_df['AR_fps'].copy()
    
    # Convert the column to dataframe with unique values and their count
    fps_unique_count_df = fps_col_series.value_counts().sort_index().to_frame()
    fps_unique_count_df.rename_axis('FPS unique values', inplace=True)
    fps_unique_count_df.rename(columns={'AR_fps':'count'})
    fps_unique_count_df['% of FPS value Distribution'] = round(fps_col_series.value_counts(normalize=True)*100, 2)
    fps_unique_count_df.sort_index()
    fps_unique_count_df.to_excel(writer, sheet_name='FPS_Distribution', index=True)
        
    # Get current sheet pointer for future writing
    curr_sheet = writer.sheets['FPS_Distribution']
    
    # Add grand total of runs
    curr_row = len(fps_unique_count_df) + 2 # update current row val
    curr_sheet.cell(row=curr_row, column=1).value = 'Grand Total'
    curr_sheet.cell(row=curr_row, column=2).value = fps_unique_count_df.sum()[0]
    curr_sheet.cell(row=curr_row, column=3).value = round(fps_unique_count_df['% of FPS value Distribution'].sum())
    
    
    # Add 3D pie chart image on the excel sheet
    data = fps_unique_count_df['% of FPS value Distribution'].values.tolist()
    labels = fps_unique_count_df.index.values.tolist()
    plt.title("Distribution of Frame Delay values, in %'")
    patches = plt.pie(data, labels=labels, autopct='%1.1f%%', startangle=120)
    plt.legend(labels, loc=5)
    piefile = f"{final_excel_file}_FPS_Distribution.png"
    plt.savefig(piefile, dpi = 100)
    img = openpyxl.drawing.image.Image(piefile)
    img.anchor = 'G4'
    curr_sheet.add_image(img)
    
    plt.close('all')
    print(f"Saved pie chart: {piefile}")


# # MAIN

# In[8]:


#main
if is_interactive():
    input_excel = 'input/consolidation_result_ARGlass_TypeA.xlsx'
else:
    input_excel = sys.argv[1]

# get the name of input excel file, discard the extension
input_excel_name, _ = os.path.splitext(os.path.basename(input_excel))

# Create output prerequisites.
#1. check if output dir exists, if not create
output_dir = 'output'
if not os.path.isdir(output_dir):
    os.mkdir(output_dir)
# Create output file name 
output_file_name = f'{input_excel_name}_post_analysis.xlsx'
# Create output file path
final_excel_file = os.path.join(output_dir,output_file_name)

# Create ExcelWriter object to populate output excel file
writer = pd.ExcelWriter(final_excel_file, engine='openpyxl')

print(f"*** Working on folder: {input_excel} ***")

# Get the input excel sheet into a dataframe
c_df = pd.read_excel(input_excel, 0, index_col=None)


# In[9]:


##### Add required sheets #######

# Sheet 1 - all runs
print("Working on Sheet 1 - All runs")
add_all_runs_sheet(c_df, writer)
print(f"Total runs: {len(c_df)} ")
print("DONE!\n")

# Sheet 2 - only non-outliers from all runs
print("Working on Sheet 2 - only non-outliers from all runs")
no_outliers_df = add_runs_without_outliers(c_df, writer)
print(f"Total non-outlier runs: {len(no_outliers_df)} ")
print("DONE!\n")

# Sheet 3 - only outliers from all runs
print("Working on Sheet 3 - only ourtliers from all runs")
with_outliers_df = add_outlier_runs(c_df, writer)
print(f"Total outlier runs: {len(with_outliers_df)} ")
print("DONE!\n")

# Sheet 4 - add statistics (min, max, stddev, etc.) for frame delay (runs without outliers)
print("Working on Sheet 4 - add statistics (min, max, stddev, etc.) for 'AR_fps' values for non-outlier runs")
fps_stat_df = add_statistics(no_outliers_df, writer)
print("DONE!\n")

# Sheet 5 - Analyze frame delay column from all runs
print("Working on Sheet 5 - Analyze frame delay column from all runs")
fps_all_analysis(c_df, writer)
print("DONE!\n")

# Final step. Save the Excel writer object and close it
print(f"Consolidating all sheets in final Excel: {final_excel_file}")
writer.save()
writer.close()
print("DONE!")


# In[ ]:




