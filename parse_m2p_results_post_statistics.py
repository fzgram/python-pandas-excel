#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import os
import sys
import csv
import re
import matplotlib.pyplot as plt
import openpyxl


# In[2]:


def is_interactive():
    import __main__ as main
    return not hasattr(main, '__file__')


# In[3]:


def get_run_number(filename):
    m = re.search('RUN(\d{1,9})', filename, re.IGNORECASE)
    if m:
        return int(m.group(1))
    else:
        print("No RUN number found. Using 'None'")
        return None


# In[4]:


def get_consolidated_df(top_folder):
    all_csv_list = []
    consolidated_df = pd.DataFrame()
    
    for root, dirnames, filenames in os.walk(top_folder):
        for files in filenames:
            if '.csv' in files:
                all_csv_list.append(os.path.join(root, files))
    
    # Sort csv's as per runs
    all_csv_list = sorted(all_csv_list, key=lambda i: int(get_run_number(i)))
    #print(all_csv_list)
    
    for csv_file in all_csv_list:
        df = pd.read_csv(csv_file)
        df.insert(0, 'run', get_run_number(csv_file))
        consolidated_df = consolidated_df.append(df)
    
    return consolidated_df
        


# ## Sheet 1 - all runs

# In[5]:


def add_all_runs_sheet(c_df, writer):
    c_df.to_excel(writer, sheet_name='All runs', index=False)    


# ## Sheet 2 - only non-outliers from all runs

# In[6]:


def add_runs_without_outliers(c_df, writer):
    no_outliers_df = c_df[c_df['frame delay @ 240 fps'] <= 4]
    no_outliers_df.to_excel(writer, sheet_name='Runs without outliers', index=False)
    return no_outliers_df


# ## Sheet 3 - only ourtliers from all runs

# In[7]:


def add_outlier_runs(c_df, writer):
    with_outliers_df = c_df[c_df['frame delay @ 240 fps'] > 4]
    with_outliers_df.to_excel(writer, sheet_name='Outlier runs', index=False)
    return with_outliers_df


# ## Sheet 4 - add statistics (min, max, stddev, etc.) for frame delay and m2p_latency values (runs without outliers)

# In[8]:


def add_statistics(c_df, writer):
    sliced_df = c_df[['run', 'frame delay @ 240 fps', 'm2p_latency']].copy()
    desc_df = sliced_df.describe()
    median_series = sliced_df.median()
    
    frame_delay_stat_df = pd.DataFrame(columns=['Min', 'Max', 'Average', 'Median', 'Std Deviation'])
    latency_stat_df = pd.DataFrame(columns=['Min', 'Max', 'Average', 'Median', 'Std Deviation'])
      
    frame_delay_stat_df['Min'] = [desc_df["frame delay @ 240 fps"]["min"]]
    frame_delay_stat_df['Max'] = [desc_df["frame delay @ 240 fps"]["max"]]
    frame_delay_stat_df['Average'] = [desc_df["frame delay @ 240 fps"]["mean"]]
    frame_delay_stat_df['Median'] = [median_series["frame delay @ 240 fps"]]
    frame_delay_stat_df['Std Deviation'] = [desc_df["frame delay @ 240 fps"]["std"]]

    latency_stat_df['Min'] = [desc_df["m2p_latency"]["min"]]
    latency_stat_df['Max'] = [desc_df["m2p_latency"]["max"]]
    latency_stat_df['Average'] = [desc_df["m2p_latency"]["mean"]]
    latency_stat_df['Median'] = [median_series["m2p_latency"]]
    latency_stat_df['Std Deviation'] = [desc_df["m2p_latency"]["std"]]
    
    # Round to 4 decimals
    frame_delay_stat_df = frame_delay_stat_df.round(4)
    latency_stat_df = latency_stat_df.round(4)
    
    # Write the sliced_df to excel
    sliced_df.to_excel(writer, sheet_name='Statistics', index=False)
    
    #get a pointer to the same sheet to write other dfs and text to the same sheet
    curr_sheet = writer.sheets['Statistics']
    
    # Write text and frame_delay_stat_df
    #curr_sheet.write(1, 4, "Statistics, # of Frames Delay")
    curr_sheet['E2'] = "Statistics, # of Frames Delay"
    frame_delay_stat_df.to_excel(writer, startrow=2, startcol=4, sheet_name='Statistics', index=False)
    
    # Write text and frame_delay_stat_df
    #curr_sheet.write(5, 4, "Statistics, Motion to Photon Latency in msec")
    curr_sheet['E6'] = "Statistics, Motion to Photon Latency in msec"
    latency_stat_df.to_excel(writer, startrow=6, startcol=4, sheet_name='Statistics', index=False)
    
    return frame_delay_stat_df, latency_stat_df


# ## Sheet 5 - Analyze frame delay column from all runs

# In[9]:


def frame_delay_all_analysis(c_df, writer):
    curr_row = 0
    frame_delay_col_series = c_df['frame delay @ 240 fps'].copy()
        
    # Convert the column to dataframe with unique values and their count
    frame_delay_unique_count_df = frame_delay_col_series.value_counts().sort_index().to_frame('Count of frame delay @ 240 fps')
    frame_delay_unique_count_df.rename_axis('Frame Delay', inplace=True)
    frame_delay_unique_count_df.to_excel(writer, sheet_name='Frame Delay-Allruns', index=True)
    #unique_counts_df = frame_delay_col_series.value_counts().sort_index().to_frame(name='count of frame delay @ 240 fps')
    #unique_counts_df.to_excel(writer, sheet_name='Frame Delay-Allruns', index=True)
    
    # Get current sheet pointer for future writing
    curr_sheet = writer.sheets['Frame Delay-Allruns']
    
    # Add grand total of runs
    curr_row = len(frame_delay_unique_count_df) + 2 # update current row val. +2 because, 1 for column names of the dataframe and 1 for next empty row
    curr_sheet.cell(row=curr_row, column=1).value = 'Grand Total'
    curr_sheet.cell(row=curr_row, column=2).value = frame_delay_unique_count_df.sum()[0]
    
    # Add outlier vs non-outlier stats
    curr_row = curr_row + 1 # update current row val
    
    outlier_stat_df = pd.DataFrame(columns=['Frames Delayed', 'Distribution of Frame Delay with all runs', 'in %'])
    total_non_outliers = len(frame_delay_col_series[frame_delay_col_series[0] <= 4])
    total_outliers = len(frame_delay_col_series[frame_delay_col_series[0] > 4])
    total_sum = frame_delay_unique_count_df.sum()[0]

    outlier_stat_df.loc[0] = ['<=4', total_non_outliers, round((total_non_outliers*100)/total_sum, 2) ]
    outlier_stat_df.loc[1] = ['>4', total_outliers, round((total_outliers*100)/total_sum, 2) ]
    outlier_stat_df.loc[2] = ['total', total_sum, 100]
    outlier_stat_df.to_excel(writer, sheet_name='Frame Delay-Allruns', startrow=curr_row, index=False)
    
    # Add 3D pie chart image on the excel sheet
    data = outlier_stat_df['in %'].values.tolist()[:-1]
    labels = ['<=4', '>4']
    plt.title("Distribution of Frame Delay with all runs, in %'")
    plt.pie(data, labels=labels, autopct='%1.1f%%', startangle=120, shadow=True)
    piefile = f"{top_folder_final_name}_FrameDelayAnalysis.png"
    plt.savefig(piefile, dpi = 75)
    img = openpyxl.drawing.image.Image(piefile)
    img.anchor = 'G4'
    curr_sheet.add_image(img)
    
    #Append piechart filename to global list
    pie_chart_files.append(piefile)
    plt.close('all')
    print(f"Saved pie chart: {piefile}")


# ## Sheet 6 - Analyze frame delay column from all non-outlier runs

# In[10]:


def frame_delay_no_outliers_analysis(no_outliers_df, writer):
    curr_row = 0
    frame_delay_col_series = no_outliers_df['frame delay @ 240 fps'].copy()
    
    # Convert the column to dataframe with unique values and their count
    frame_delay_unique_count_df = frame_delay_col_series.value_counts().sort_index().to_frame()
    frame_delay_unique_count_df.rename_axis('Frames Delayed', inplace=True)
    frame_delay_unique_count_df['% of Frame Delay Distribution'] = round(frame_delay_col_series.value_counts(normalize=True)*100, 2)
    frame_delay_unique_count_df.sort_index()
    frame_delay_unique_count_df.to_excel(writer, sheet_name='Frame Delay-Runswithoutoutliers', index=True)
        
    # Get current sheet pointer for future writing
    curr_sheet = writer.sheets['Frame Delay-Runswithoutoutliers']
    
    # Add grand total of runs
    curr_row = len(frame_delay_unique_count_df) + 2 # update current row val
    curr_sheet.cell(row=curr_row, column=1).value = 'Grand Total'
    curr_sheet.cell(row=curr_row, column=2).value = frame_delay_unique_count_df.sum()[0]
    curr_sheet.cell(row=curr_row, column=3).value = round(frame_delay_unique_count_df['% of Frame Delay Distribution'].sum())
    
    
    # Add 3D pie chart image on the excel sheet
    data = frame_delay_unique_count_df['% of Frame Delay Distribution'].values.tolist()
    labels = frame_delay_unique_count_df.index.values.tolist()
    plt.title("Distribution of Frame Delay with non-outlier runs, in %'")
    plt.pie(data, labels=labels, autopct='%1.1f%%', startangle=120)
    piefile = f"{top_folder_final_name}_FrameDelayAnalysisNonOutliers.png"
    plt.savefig(piefile, dpi = 75)
    img = openpyxl.drawing.image.Image(piefile)
    img.anchor = 'G4'
    curr_sheet.add_image(img)
    
    #Append piechart filename to global list
    pie_chart_files.append(piefile)
    plt.close('all')
    print(f"Saved pie chart: {piefile}")


# ## Sheet 7 - Analyze m2p_latency distribution of non-outlier runs

# In[11]:


def latency_distribution(no_outliers_df, writer):
    curr_row = 0
    m2p_latency_col_series = no_outliers_df['m2p_latency'].copy()
    
    # Convert the column to dataframe with unique values and their count
    m2p_latency_unique_count_df = m2p_latency_col_series.value_counts().sort_index().to_frame()
    m2p_latency_unique_count_df.rename_axis('Frame delay', inplace=True)
    m2p_latency_unique_count_df['% of Motion to Photon Latency Distribution'] = round(m2p_latency_col_series.value_counts(normalize=True)*100, 2)
    m2p_latency_unique_count_df.sort_index()
    m2p_latency_unique_count_df.to_excel(writer, sheet_name='Distribution-Latency(msec)', index=True)
        
    # Get current sheet pointer for future writing
    curr_sheet = writer.sheets['Distribution-Latency(msec)']
    
    # Add grand total of runs
    curr_row = len(m2p_latency_unique_count_df) + 2 # update current row val
    curr_sheet.cell(row=curr_row, column=1).value = 'Grand Total'
    curr_sheet.cell(row=curr_row, column=2).value = m2p_latency_unique_count_df.sum()[0]
    curr_sheet.cell(row=curr_row, column=3).value = round(m2p_latency_unique_count_df['% of Motion to Photon Latency Distribution'].sum())
    
    # Add 3D pie chart image on the excel sheet
    data = m2p_latency_unique_count_df['% of Motion to Photon Latency Distribution'].values.tolist()
    labels = m2p_latency_unique_count_df.index.values.tolist()
    plt.title("Distribution of Motion to Photon Latency (msec)")
    plt.pie(data, labels=labels, autopct='%1.1f%%', startangle=120)
    piefile = f"{top_folder_final_name}_M2PLatencyDistributionNonOutliers.png"
    plt.savefig(piefile, dpi = 75)
    img = openpyxl.drawing.image.Image(piefile)
    img.anchor = 'G4'
    curr_sheet.add_image(img)
    
    #Append piechart filename to global list
    pie_chart_files.append(piefile)
    plt.close('all')
    print(f"Saved pie chart: {piefile}")


# ## Sheet 8 - combine all pie charts from Sheet 5,6,7

# In[12]:


def combine_all_piecharts(writer):
    df = pd.DataFrame(columns=['Total charts'])
    df.loc[0] = len(pie_chart_files)
    df.to_excel(writer, sheet_name='Combined_PieCharts', index=False)
    curr_sheet = writer.sheets['Combined_PieCharts']
    
    col = 'A'
    row = 4
    for piefile in pie_chart_files:
        img = openpyxl.drawing.image.Image(piefile)
        img.anchor = f'{col}{row}'
        curr_sheet.add_image(img)
        row = row + 18


# # MAIN

# In[14]:


#main
if is_interactive():
    top_folder = '07012019_roman_content'
else:
    top_folder = sys.argv[1]
    
top_folder_final_name = top_folder.replace('/' , '_')
excel_file = f'consolidation_result_{top_folder_final_name}.xlsx'
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

print(f"*** Working on folder: {top_folder} ***")

# list of names of all piecharts to create the last excel sheet
pie_chart_files = []

print(f"Consolidating results of multiples runs in {top_folder}")
c_df = get_consolidated_df(top_folder)
print("DONE!\n")

##### Add required sheets #######

# Sheet 1 - all runs
print("Working on Sheet 1 - All runs")
add_all_runs_sheet(c_df, writer)
print(f"Total runs: {len(c_df)} ")
print("DONE!\n")

# Sheet 2 - only non-outliers from all runs
print("Working on Sheet 2 - only non-outliers from all runs")
no_outliers_df = add_runs_without_outliers(c_df, writer)
print(f"Total non-outlier runs: {len(no_outliers_df)} ")
print("DONE!\n")

# Sheet 3 - only ourtliers from all runs
print("Working on Sheet 3 - only ourtliers from all runs")
with_outliers_df = add_outlier_runs(c_df, writer)
print(f"Total outlier runs: {len(with_outliers_df)} ")
print("DONE!\n")

# Sheet 4 - add statistics (min, max, stddev, etc.) for frame delay and m2P_latency values (runs without outliers)
print("Working on Sheet 4 - add statistics (min, max, stddev, etc.) for 'frame delay' and 'M2P_latency' values for non-outlier runs")
frame_delay_stat_df, latency_stat_df = add_statistics(no_outliers_df, writer)
print("DONE!\n")

# Sheet 5 - Analyze frame delay column from all runs
print("Working on Sheet 5 - Analyze frame delay column from all runs")
frame_delay_all_analysis(c_df, writer)
print("DONE!\n")

# Sheet 6 - Analyze frame delay distribution of non-outlier runs
print("Working on Sheet 6 - Analyze frame delay distribution of non-outlier runs")
frame_delay_no_outliers_analysis(no_outliers_df, writer)
print("DONE!\n")

# Sheet 7 - Analyze m2p_latency distribution of non-outlier runs
print("Working on Sheet 7 - Analyze m2p_latency distribution of non-outlier runs")
latency_distribution(no_outliers_df, writer)
print("DONE!\n")

# Sheet 8 - combine all pie charts from Sheet 5,6,7
print("Working on Sheet 8 - combine all pie charts from Sheet 5,6,7")
combine_all_piecharts(writer)
print("DONE!\n")

# Final step. Sabe the Excel writer object and close it
print(f"Consolidating all sheets in final Excel: {excel_file}")
writer.save()
writer.close()
print("DONE!")


# In[ ]:




